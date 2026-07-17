"""Lecture tabulaire (CSV / XLSX) et normalisation de valeurs."""
from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path


def read_tabular(path: str | Path) -> tuple[list[str], list[dict]]:
    """Retourne (colonnes, lignes) depuis un fichier CSV ou XLSX."""
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Format non pris en charge : {suffix}")


def _read_csv(path: str | Path) -> tuple[list[str], list[dict]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=dialect)
        columns = [c.strip() for c in (reader.fieldnames or [])]
        rows = [
            {(k or "").strip(): (v.strip() if isinstance(v, str) else v)
             for k, v in row.items()}
            for row in reader
        ]
    return columns, rows


def _read_xlsx(path: str | Path) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None) or ()
        columns = [str(c).strip() if c is not None else "" for c in header]
        rows: list[dict] = []
        for raw in rows_iter:
            if raw is None or all(c is None for c in raw):
                continue
            row = {}
            for idx, col in enumerate(columns):
                if not col:
                    continue
                value = raw[idx] if idx < len(raw) else None
                row[col] = value.strip() if isinstance(value, str) else value
            rows.append(row)
    finally:
        wb.close()
    return columns, rows


# -------- normaliseurs --------
def norm_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def norm_float(value) -> float | None:
    text = norm_str(value)
    if text is None:
        return None
    text = text.replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Nombre invalide : « {value} »") from exc


def norm_int(value) -> int | None:
    f = norm_float(value)
    return int(f) if f is not None else None


def norm_bool(value) -> bool:
    text = (norm_str(value) or "").lower()
    return text in {"1", "true", "vrai", "oui", "yes", "x"}


def norm_keywords(value) -> list[str]:
    text = norm_str(value)
    if not text:
        return []
    parts = [p.strip() for p in text.replace(";", ",").split(",")]
    return [p for p in parts if p]


def norm_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide : « {value} »")


def norm_datetime(value) -> datetime | None:
    d = norm_date(value)
    return datetime(d.year, d.month, d.day) if d else None

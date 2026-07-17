"""Exports du rapport de conformité : CSV, Excel, PDF (repli HTML)."""
from __future__ import annotations

import csv
import io

_ITEM_HEADERS = [
    ("line", "Ligne"),
    ("raw_name", "Produit (facture)"),
    ("matched_name", "Produit rapproché"),
    ("confidence", "Confiance"),
    ("hs_code", "Code SH"),
    ("import_duty", "Droit import. %"),
    ("vat", "TVA %"),
    ("status", "Statut"),
    ("invoice_price", "Prix facturé"),
    ("average_price", "Prix moyen"),
    ("price_variation_percent", "Variation %"),
]


def to_csv(content: dict) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in _ITEM_HEADERS])
    for item in content.get("detected_products", []):
        writer.writerow([_cell(item.get(key)) for key, _ in _ITEM_HEADERS])
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx(content: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    inv = content.get("invoice_summary", {})
    ws.append(["Rapport de conformité à l'import — UM6P"])
    ws.append(["Facture", inv.get("invoice_number") or "—"])
    ws.append(["Date", inv.get("invoice_date") or "—"])
    ws.append(["Devise", inv.get("currency") or "—"])
    ws.append(["Incoterm", inv.get("incoterm") or "—"])
    ws.append(["Risque global", content.get("compliance_analysis", {}).get("overall_risk", "—")])
    ws.append(["Confiance", content.get("confidence", "—")])

    items_ws = wb.create_sheet("Produits")
    items_ws.append([label for _, label in _ITEM_HEADERS])
    for item in content.get("detected_products", []):
        items_ws.append([_cell(item.get(key)) for key, _ in _ITEM_HEADERS])

    findings_ws = wb.create_sheet("Conformité")
    findings_ws.append(["Type", "Risque", "Message"])
    for f in content.get("compliance_analysis", {}).get("findings", []):
        findings_ws.append([f.get("type"), f.get("risk"), f.get("message")])

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def build_html(content: dict) -> str:
    inv = content.get("invoice_summary", {})
    comp = content.get("compliance_analysis", {})
    rows = "".join(
        "<tr>"
        + "".join(f"<td>{_cell(item.get(key))}</td>" for key, _ in _ITEM_HEADERS)
        + "</tr>"
        for item in content.get("detected_products", [])
    )
    warnings = "".join(
        f"<li><b>{w.get('risk')}</b> — {w.get('message')}</li>"
        for w in content.get("warnings", [])
    )
    recos = "".join(f"<li>{r}</li>" for r in content.get("recommendations", []))
    sources = ", ".join(content.get("sources", [])) or "—"
    headers = "".join(f"<th>{label}</th>" for _, label in _ITEM_HEADERS)
    return f"""<!doctype html><html lang="fr"><head><meta charset="utf-8">
<title>Rapport de conformité UM6P</title>
<style>
body{{font-family:system-ui,Arial,sans-serif;color:#3B3B3C;margin:32px;}}
h1{{color:#D7492A;font-size:20px;}} h2{{font-size:14px;border-bottom:2px solid #ED6E47;padding-bottom:4px;}}
table{{border-collapse:collapse;width:100%;font-size:12px;}} th,td{{border:1px solid #ddd;padding:6px;text-align:left;}}
th{{background:#faf3f0;}} .risk{{font-weight:bold;color:#D7492A;}}
</style></head><body>
<h1>Rapport de conformité à l'import — UM6P</h1>
<p>Facture <b>{inv.get('invoice_number') or '—'}</b> · {inv.get('invoice_date') or '—'}
· {inv.get('currency') or '—'} · Incoterm {inv.get('incoterm') or '—'}</p>
<p>Risque global : <span class="risk">{comp.get('overall_risk','—')}</span> ·
Confiance : {content.get('confidence','—')}</p>
<h2>Produits détectés</h2>
<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>
<h2>Avertissements</h2><ul>{warnings or '<li>Aucun</li>'}</ul>
<h2>Recommandations</h2><ul>{recos or '<li>Aucune</li>'}</ul>
<h2>Sources</h2><p>{sources}</p>
</body></html>"""


def build_pdf(content: dict) -> bytes | None:
    """Rapport PDF « entreprise » via reportlab ; None si indisponible."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return None

    stream = io.BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4, title="Rapport de conformité UM6P")
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#D7492A")
    story = []

    inv = content.get("invoice_summary", {})
    comp = content.get("compliance_analysis", {})

    title = styles["Title"]
    title.textColor = brand
    story.append(Paragraph("Rapport de conformité à l'import — UM6P", title))
    story.append(Paragraph(
        f"Facture {inv.get('invoice_number') or '—'} · {inv.get('invoice_date') or '—'} · "
        f"{inv.get('currency') or '—'} · Incoterm {inv.get('incoterm') or '—'}",
        styles["Normal"],
    ))
    story.append(Paragraph(
        f"<b>Risque global :</b> {comp.get('overall_risk', '—')} · "
        f"<b>Confiance :</b> {content.get('confidence', '—')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 6 * mm))

    data = [[label for _, label in _ITEM_HEADERS]]
    for item in content.get("detected_products", []):
        data.append([str(_cell(item.get(key))) for key, _ in _ITEM_HEADERS])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#faf3f0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), brand),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(Paragraph("Produits détectés", styles["Heading2"]))
    story.append(table)
    story.append(Spacer(1, 6 * mm))

    story.append(Paragraph("Avertissements", styles["Heading2"]))
    for w in content.get("warnings", []) or [{"message": "Aucun"}]:
        story.append(Paragraph(f"• [{w.get('risk', '')}] {w.get('message')}", styles["Normal"]))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Recommandations", styles["Heading2"]))
    for r in content.get("recommendations", []) or ["Aucune"]:
        story.append(Paragraph(f"• {r}", styles["Normal"]))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "Sources : " + (", ".join(content.get("sources", [])) or "—"), styles["Italic"]
    ))

    doc.build(story)
    return stream.getvalue()


def _cell(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)

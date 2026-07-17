"""Extraction XLSX (openpyxl). Une « page » logique = une feuille."""
from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from app.ingestion.extractors.base import CorruptedDocumentError, Extractor
from app.ingestion.types import ExtractedPage


class XlsxExtractor(Extractor):
    extensions = {".xlsx"}

    def _open(self, path: str | Path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl n'est pas installé.") from exc
        try:
            return load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise CorruptedDocumentError(f"XLSX illisible : {exc}") from exc

    def count_pages(self, path: str | Path) -> int:
        wb = self._open(path)
        try:
            return len(wb.sheetnames)
        finally:
            wb.close()

    def iter_pages(self, path: str | Path) -> Iterator[ExtractedPage]:
        wb = self._open(path)
        try:
            for index, sheet in enumerate(wb.worksheets, start=1):
                lines: list[str] = [f"[Feuille] {sheet.title}"]
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append(" | ".join(cells))
                yield ExtractedPage(number=index, text="\n".join(lines))
        finally:
            wb.close()
